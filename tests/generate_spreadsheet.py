"""Generate an Excel spreadsheet from load test results.

Usage:
    python tests/generate_spreadsheet.py --summary reports/load_test_XYZ/summary.json \
        --locust-csv reports/load_test_XYZ/stats_stats.csv \
        --output reports/load_test_XYZ/load_test_results.xlsx
"""

import argparse
import csv
import json
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Styles ───────────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

GREEN_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
PASS_FONT = Font(name="Calibri", bold=True, color="059669")
WARN_FONT = Font(name="Calibri", bold=True, color="D97706")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

NUM_FMT_2DP = "0.00"
NUM_FMT_DOLLAR = "$#,##0.00"
NUM_FMT_COMMA = "#,##0"
NUM_FMT_PCT = "0.00%"


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_data_cell(cell):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def latency_fill(ms):
    if ms < 200:
        return GREEN_FILL
    if ms < 1000:
        return YELLOW_FILL
    return RED_FILL


# ═════════════════════════════════════════════════════════════════════════════
#  Sheet builders
# ═════════════════════════════════════════════════════════════════════════════

def build_executive_summary(wb, tiers):
    ws = wb.active
    ws.title = "Executive Summary"

    headers = ["Tier", "Target RPS", "Actual RPS", "Achievement %",
               "Total Requests", "Errors", "Error Rate %", "P50 (ms)",
               "P90 (ms)", "P95 (ms)", "P99 (ms)", "Status"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    for i, t in enumerate(tiers):
        row = i + 2
        pct = round(t["actual_rps"] / t["target_rps"] * 100, 1) if t["target_rps"] > 0 else 0
        passed = t["error_rate"] < 5 and t["actual_rps"] >= t["target_rps"] * 0.8
        status = "PASS" if passed else "WARN"

        values = [
            t["tier"], t["target_rps"], t["actual_rps"], pct / 100,
            t["total_requests"], t["error_count"], t["error_rate"] / 100,
            t["latency"]["p50"], t["latency"]["p90"],
            t["latency"]["p95"], t["latency"]["p99"], status,
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            style_data_cell(cell)

        # Number formats
        ws.cell(row=row, column=4).number_format = NUM_FMT_PCT
        ws.cell(row=row, column=7).number_format = NUM_FMT_PCT
        for col in (8, 9, 10, 11):
            ws.cell(row=row, column=col).number_format = NUM_FMT_2DP

        # Status coloring
        status_cell = ws.cell(row=row, column=12)
        status_cell.font = PASS_FONT if passed else WARN_FONT

    # Throughput chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Target vs Actual RPS by Tier"
    chart.y_axis.title = "Requests per Second"
    chart.x_axis.title = "Tier"
    chart.style = 10
    data_ref = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=len(tiers) + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(tiers) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 18
    chart.height = 12
    ws.add_chart(chart, f"A{len(tiers) + 4}")

    auto_width(ws)


def build_latency_by_tier(wb, tiers):
    ws = wb.create_sheet("Latency by Tier")

    row = 1
    for t in tiers:
        # Tier subheader
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=f"Tier {t['tier']} — {t['target_rps']} RPS target ({t['actual_rps']} actual)")
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="left")
        row += 1

        headers = ["Endpoint", "Count", "Error Count", "Error %", "P50 (ms)", "P90 (ms)", "P95 (ms)", "P99 (ms)"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        for ep_name in sorted(t["per_endpoint"].keys()):
            stats = t["per_endpoint"][ep_name]
            values = [
                ep_name, stats["count"], stats["error_count"],
                stats["error_rate"] / 100,
                stats["p50"], stats["p90"], stats["p95"], stats["p99"],
            ]
            for c, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=c, value=v)
                style_data_cell(cell)
                if c == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            ws.cell(row=row, column=4).number_format = NUM_FMT_PCT
            for col in (5, 6, 7, 8):
                cell = ws.cell(row=row, column=col)
                cell.number_format = NUM_FMT_2DP
                cell.fill = latency_fill(cell.value)

            row += 1

        row += 1  # blank row between tiers

    auto_width(ws)


def build_error_analysis(wb, tiers):
    ws = wb.create_sheet("Error Analysis")

    headers = ["Tier", "Target RPS", "Total Requests", "Errors", "Error Rate %"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    for i, t in enumerate(tiers):
        row = i + 2
        values = [t["tier"], t["target_rps"], t["total_requests"],
                  t["error_count"], t["error_rate"] / 100]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            style_data_cell(cell)
        ws.cell(row=row, column=3).number_format = NUM_FMT_COMMA
        ws.cell(row=row, column=5).number_format = NUM_FMT_PCT

    # Per-tier per-endpoint errors
    row = len(tiers) + 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="Endpoints with Errors (>0%)").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    row += 1

    err_headers = ["Tier", "Endpoint", "Count", "Errors", "Error Rate %"]
    for c, h in enumerate(err_headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(err_headers))
    row += 1

    for t in tiers:
        for ep_name in sorted(t["per_endpoint"].keys()):
            stats = t["per_endpoint"][ep_name]
            if stats["error_count"] > 0:
                values = [t["tier"], ep_name, stats["count"],
                          stats["error_count"], stats["error_rate"] / 100]
                for c, v in enumerate(values, 1):
                    cell = ws.cell(row=row, column=c, value=v)
                    style_data_cell(cell)
                    if c == 2:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                ws.cell(row=row, column=5).number_format = NUM_FMT_PCT
                row += 1

    auto_width(ws)


def build_bottlenecks(wb, bottlenecks):
    ws = wb.create_sheet("Bottlenecks")

    headers = ["Tier", "Type", "Endpoint", "Detail", "Recommendation"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    for i, b in enumerate(bottlenecks):
        row = i + 2
        values = [str(b["tier"]), b["type"], b["endpoint"], b["detail"], b["recommendation"]]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            style_data_cell(cell)
            if c >= 4:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    auto_width(ws)
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 55


def build_cost_projections(wb, cost_projections):
    ws = wb.create_sheet("AWS Cost Projection")

    headers = ["Tier", "Target RPS", "Actual RPS", "Scenario",
               "Monthly Requests", "Lambda Invocation ($)", "Lambda Duration ($)",
               "API Gateway ($)", "DynamoDB ($)", "Total Monthly ($)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    for i, cp in enumerate(cost_projections):
        row = i + 2
        values = [
            cp["tier"], cp["target_rps"], cp["actual_rps"], cp["scenario"],
            cp["monthly_requests"], cp["lambda_invocation"], cp["lambda_duration"],
            cp["api_gateway"], cp["dynamodb_estimated"], cp["total_monthly"],
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            style_data_cell(cell)
            if c == 4:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.cell(row=row, column=5).number_format = NUM_FMT_COMMA
        for col in (6, 7, 8, 9, 10):
            ws.cell(row=row, column=col).number_format = NUM_FMT_DOLLAR

    # Total cost chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Cost by Tier & Scenario"
    chart.y_axis.title = "USD / month"
    chart.style = 10
    total_col = 10
    data_ref = Reference(ws, min_col=total_col, min_row=1, max_row=len(cost_projections) + 1)
    cats_labels = []
    for cp in cost_projections:
        cats_labels.append(f"T{cp['tier']} {cp['scenario'][:5]}")
    chart.add_data(data_ref, titles_from_data=True)
    # Build category labels
    label_col = len(headers) + 2
    for i, label in enumerate(cats_labels):
        ws.cell(row=i + 2, column=label_col, value=label)
    cats = Reference(ws, min_col=label_col, min_row=2, max_row=len(cost_projections) + 1)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 24
    chart.height = 14
    ws.add_chart(chart, f"A{len(cost_projections) + 4}")

    auto_width(ws)


def build_locust_raw_stats(wb, locust_csv_path):
    ws = wb.create_sheet("Locust Raw Stats")

    if not os.path.exists(locust_csv_path):
        ws.cell(row=1, column=1, value="Locust CSV not found")
        return

    with open(locust_csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for r, csv_row in enumerate(reader, 1):
            for c, val in enumerate(csv_row, 1):
                cell = ws.cell(row=r, column=c)
                # Try to convert numeric values
                try:
                    cell.value = float(val)
                    cell.number_format = NUM_FMT_2DP
                except ValueError:
                    cell.value = val
                style_data_cell(cell)

    style_header_row(ws, 1, ws.max_column)
    auto_width(ws)


def build_diagnostics(wb, diagnostics):
    ws = wb.create_sheet("Diagnostics")

    # Analytics summary
    row = 1
    ws.cell(row=row, column=1, value="Analytics Summary").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    analytics = diagnostics.get("analytics_summary", {})
    if "error" not in analytics:
        for key in ("total_orders", "accepted_items", "rejected_items", "acceptance_rate"):
            ws.cell(row=row, column=1, value=key.replace("_", " ").title())
            ws.cell(row=row, column=2, value=analytics.get(key, "N/A"))
            for c in (1, 2):
                style_data_cell(ws.cell(row=row, column=c))
                ws.cell(row=row, column=c).alignment = Alignment(horizontal="left")
            row += 1

    row += 1

    # Top SKUs
    top_skus = diagnostics.get("top_skus", {})
    if "error" not in top_skus:
        sku_list = top_skus.get("top_skus", [])
        if sku_list:
            ws.cell(row=row, column=1, value="Top SKUs").font = SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = SUBHEADER_FILL
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            row += 1
            for c, h in enumerate(["Rank", "SKU", "Total Quantity"], 1):
                ws.cell(row=row, column=c, value=h)
            style_header_row(ws, row, 3)
            row += 1
            for i, s in enumerate(sku_list[:10], 1):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=s.get("sku", ""))
                ws.cell(row=row, column=3, value=s.get("total_quantity", ""))
                for c in (1, 2, 3):
                    style_data_cell(ws.cell(row=row, column=c))
                row += 1

    row += 1

    # Saga state distribution
    sagas = diagnostics.get("sagas", {})
    if isinstance(sagas, dict) and "error" not in sagas:
        saga_list = sagas.get("sagas", [])
        if saga_list:
            status_counts = {}
            for s in saga_list:
                st = s.get("current_status", "unknown")
                status_counts[st] = status_counts.get(st, 0) + 1

            ws.cell(row=row, column=1, value="Saga State Distribution").font = SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = SUBHEADER_FILL
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1
            for c, h in enumerate(["Status", "Count"], 1):
                ws.cell(row=row, column=c, value=h)
            style_header_row(ws, row, 2)
            row += 1
            for st, cnt in sorted(status_counts.items()):
                ws.cell(row=row, column=1, value=st)
                ws.cell(row=row, column=2, value=cnt)
                for c in (1, 2):
                    style_data_cell(ws.cell(row=row, column=c))
                row += 1

    auto_width(ws)


# ═════════════════════════════════════════════════════════════════════════════
#  Main
# ═════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Generate Excel spreadsheet from load test results")
    parser.add_argument("--summary", required=True, help="Path to summary.json")
    parser.add_argument("--locust-csv", required=True, help="Path to stats_stats.csv")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    with open(args.summary, encoding="utf-8") as f:
        data = json.load(f)

    tiers = data["tiers"]
    bottlenecks = data["bottlenecks"]
    cost_projections = data["cost_projections"]
    diagnostics = data["diagnostics"]

    wb = Workbook()

    print("[spreadsheet] Building Executive Summary...")
    build_executive_summary(wb, tiers)

    print("[spreadsheet] Building Latency by Tier...")
    build_latency_by_tier(wb, tiers)

    print("[spreadsheet] Building Error Analysis...")
    build_error_analysis(wb, tiers)

    print("[spreadsheet] Building Bottlenecks...")
    build_bottlenecks(wb, bottlenecks)

    print("[spreadsheet] Building AWS Cost Projection...")
    build_cost_projections(wb, cost_projections)

    print("[spreadsheet] Building Locust Raw Stats...")
    build_locust_raw_stats(wb, args.locust_csv)

    print("[spreadsheet] Building Diagnostics...")
    build_diagnostics(wb, diagnostics)

    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
    wb.save(args.output)
    print(f"\n[spreadsheet] Done! Saved to: {args.output}")


if __name__ == "__main__":
    main()
